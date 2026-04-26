import asyncio
import contextlib
import json
import os
import shutil
import socket
import subprocess
import sys
import threading
import uuid
from io import BytesIO
from datetime import datetime
from pathlib import Path
from typing import Any
from urllib.parse import quote

import psutil
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from fastapi import FastAPI, File, Form, HTTPException, Request, UploadFile
from fastapi.responses import JSONResponse, StreamingResponse
from fastapi.templating import Jinja2Templates

from app_config import APP_CONFIG

BASE_DIR = APP_CONFIG.base_dir
RUNS_DIR = APP_CONFIG.runs_dir
UPLOADS_DIR = APP_CONFIG.uploads_dir
STATE_FILE = APP_CONFIG.state_file
COOKIE_FILE = APP_CONFIG.cookie_file
OPEN_SCRIPT = BASE_DIR / 'open_pdd_goods.py'
LOGIN_SCRIPT = BASE_DIR / 'save_pdd_cookie.py'
TEMPLATES_DIR = BASE_DIR / 'templates'
EXTENSION_DIR = APP_CONFIG.extension_dir
DEFAULT_USER_DATA_DIR = APP_CONFIG.user_data_dir
BROWSER_PATH = APP_CONFIG.browser_path
MAX_PARALLEL_JOBS_HARD_LIMIT = 1

DEFAULT_SETTINGS: dict[str, Any] = {
    'max_batch_size': 50,
    'concurrency': 1,
    'review_timeout': 80,
    'default_timeout': 15,
    'poll_interval': 1,
    'plugin_ready_retry_times': 10,
    'plugin_ready_retry_interval': 2,
    'search_result_retry_times': 5,
    'search_result_retry_interval': 3,
    'bulk_action_retry_times': 5,
    'bulk_action_retry_interval': 1,
    'select_all_retry_times': 5,
    'select_all_retry_interval': 1,
    'restore_dialog_retry_times': 5,
    'restore_dialog_retry_interval': 3,
    'flash_sale_amount': '10',
    'text_to_find': '',
    'text_to_replace': '',
    'max_parallel_jobs': 1,
    'browser_profile_mode': 'shared',
}


app = FastAPI(title='拼多多本机自动化控制台')
templates = Jinja2Templates(directory=str(TEMPLATES_DIR))
state_lock = threading.Lock()
runtime_lock = threading.Lock()
job_runtimes: dict[str, dict[str, Any]] = {}
login_runtime: dict[str, Any] = {
    'status': 'idle',
    'logs': [],
    'started_at': None,
    'ended_at': None,
    'returncode': None,
}


def now_iso() -> str:
    return datetime.now().isoformat(timespec='seconds')


def default_state() -> dict[str, Any]:
    return {'last_settings': DEFAULT_SETTINGS.copy(), 'jobs': [], 'uploads': {'excel': [], 'images': []}}


def load_state() -> dict[str, Any]:
    source_file = STATE_FILE
    if not source_file.exists() and APP_CONFIG.legacy_state_file.exists():
        source_file = APP_CONFIG.legacy_state_file
    if not source_file.exists():
        return default_state()
    try:
        state = json.loads(source_file.read_text(encoding='utf-8'))
    except json.JSONDecodeError:
        return default_state()
    if not isinstance(state, dict):
        return default_state()
    state.setdefault('last_settings', DEFAULT_SETTINGS.copy())
    state.setdefault('jobs', [])
    state.setdefault('uploads', {'excel': [], 'images': []})
    state['uploads'].setdefault('excel', [])
    state['uploads'].setdefault('images', [])
    merged_settings = DEFAULT_SETTINGS.copy()
    merged_settings.update(state.get('last_settings') or {})
    merged_settings['max_parallel_jobs'] = 1
    merged_settings['browser_profile_mode'] = 'shared'
    state['last_settings'] = merged_settings
    return state


def save_state(state: dict[str, Any]) -> None:
    STATE_FILE.parent.mkdir(parents=True, exist_ok=True)
    STATE_FILE.write_text(json.dumps(state, ensure_ascii=False, indent=2), encoding='utf-8')


def update_state(mutator) -> dict[str, Any]:
    with state_lock:
        state = load_state()
        mutator(state)
        save_state(state)
        return state


def job_process_alive(job: dict[str, Any]) -> bool:
    job_id = str(job.get('job_id') or '')
    if not job_id:
        return False
    for process in psutil.process_iter(['cmdline']):
        cmdline = ' '.join(process.info.get('cmdline') or [])
        if 'open_pdd_goods.py' in cmdline and job_id in cmdline:
            return True
    return False


def reconcile_stale_jobs() -> None:
    active_statuses = {'pending', 'running', 'paused', 'stopping'}

    def mutate(state: dict[str, Any]) -> None:
        for job in state.get('jobs', []):
            status = job.get('status')
            if status not in active_statuses:
                continue
            with runtime_lock:
                tracked = job.get('job_id') in job_runtimes
            if not tracked and not job_process_alive(job):
                job['status'] = 'interrupted'
                job['ended_at'] = job.get('ended_at') or now_iso()
                job['returncode'] = job.get('returncode') if job.get('returncode') is not None else -1

    update_state(mutate)


def cookie_status() -> dict[str, Any]:
    info = {
        'exists': COOKIE_FILE.exists(),
        'path': str(COOKIE_FILE),
        'legacy_exists': APP_CONFIG.legacy_cookie_file.exists(),
        'legacy_path': str(APP_CONFIG.legacy_cookie_file),
        'saved_at': None,
        'current_url': None,
        'cookie_count': 0,
    }
    if not COOKIE_FILE.exists():
        return info
    try:
        payload = json.loads(COOKIE_FILE.read_text(encoding='utf-8'))
    except json.JSONDecodeError:
        return info
    cookies = payload.get('cookies') if isinstance(payload, dict) else None
    info['saved_at'] = payload.get('saved_at') if isinstance(payload, dict) else None
    info['current_url'] = payload.get('current_url') if isinstance(payload, dict) else None
    info['cookie_count'] = len(cookies) if isinstance(cookies, list) else 0
    return info


def extension_registered_paths(user_data_dir: Path = DEFAULT_USER_DATA_DIR) -> list[str]:
    paths: list[str] = []
    for preference_file in [
        user_data_dir / 'Default' / 'Preferences',
        user_data_dir / 'Default' / 'Secure Preferences',
    ]:
        if not preference_file.exists():
            continue
        try:
            payload = json.loads(preference_file.read_text(encoding='utf-8', errors='ignore'))
        except json.JSONDecodeError:
            continue
        settings = (((payload.get('extensions') or {}).get('settings')) or {})
        for config in settings.values():
            path_value = config.get('path')
            if isinstance(path_value, str) and path_value not in paths:
                paths.append(path_value)
    return paths


def extension_status() -> dict[str, Any]:
    expected_path = str(EXTENSION_DIR.resolve())
    registered_paths = extension_registered_paths()
    is_registered = expected_path in registered_paths
    return {
        'exists': EXTENSION_DIR.exists(),
        'path': expected_path,
        'profile_path': str(DEFAULT_USER_DATA_DIR.resolve()),
        'browser_path': str(BROWSER_PATH.resolve()) if BROWSER_PATH else '',
        'browser_exists': bool(BROWSER_PATH and BROWSER_PATH.exists()),
        'browser_version': APP_CONFIG.browser_version,
        'browser_supported': APP_CONFIG.browser_supported,
        'data_dir': str(APP_CONFIG.data_dir.resolve()),
        'cookie_path': str(COOKIE_FILE.resolve()),
        'config_file': str(APP_CONFIG.config_file.resolve()),
        'sources': APP_CONFIG.sources,
        'errors': APP_CONFIG.errors,
        'warnings': APP_CONFIG.warnings,
        'legacy_cookie_exists': APP_CONFIG.legacy_cookie_file.exists(),
        'legacy_profile_exists': APP_CONFIG.legacy_user_data_dir.exists(),
        'default_profile_registered': is_registered,
        'registered_count': len(registered_paths),
        'auto_load_enabled': True,
        'restart_before_launch_enabled': True,
        'max_parallel_jobs': 1,
    }


def automation_environment_errors() -> list[str]:
    errors = list(APP_CONFIG.errors)
    if not BROWSER_PATH or not BROWSER_PATH.exists():
        errors.append('浏览器程序不可用，请运行 prepare_deploy.py 或设置 PDD_BROWSER_PATH。')
    if not APP_CONFIG.browser_supported:
        errors.append('当前浏览器不支持自动加载扩展，请使用随包 Chromium。')
    if not EXTENSION_DIR.exists():
        errors.append('扩展目录不可用，请运行 prepare_deploy.py 或设置 PDD_EXTENSION_DIR。')
    return list(dict.fromkeys(errors))


def read_results_file(file_path: str | Path | None) -> list[dict[str, Any]]:
    if not file_path:
        return []
    path = Path(str(file_path))
    if not path.exists():
        return []
    try:
        payload = json.loads(path.read_text(encoding='utf-8'))
    except json.JSONDecodeError:
        return []
    results = payload.get('results') if isinstance(payload, dict) else None
    if not isinstance(results, list):
        return []
    return [item for item in results if isinstance(item, dict)]


def is_passed_result(item: dict[str, Any]) -> bool:
    status = str(item.get('review_status') or item.get('status') or '').strip()
    return '已通过' in status or status == '通过' or status.lower() in {'passed', 'success', 'completed'}


def result_summary_for_job(job: dict[str, Any]) -> dict[str, int]:
    paths = job.get('paths') or {}
    review_results = read_results_file(paths.get('review_results_file'))
    restore_results = read_results_file(paths.get('restore_results_file'))
    passed = [item for item in review_results if is_passed_result(item)]
    failed = [item for item in review_results if not is_passed_result(item)]
    return {
        'review_total': len(review_results),
        'review_passed': len(passed),
        'review_failed': len(failed),
        'restore_total': len(restore_results),
    }


def enrich_job(job: dict[str, Any]) -> dict[str, Any]:
    enriched = dict(job)
    enriched['result_summary'] = result_summary_for_job(job)
    return enriched


def find_job(job_id: str) -> dict[str, Any]:
    state = load_state()
    job = next((item for item in state.get('jobs', []) if item.get('job_id') == job_id), None)
    if not job:
        raise HTTPException(status_code=404, detail='任务不存在')
    return job


def classify_review_results(review_results: list[dict[str, Any]]) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    passed = [item for item in review_results if is_passed_result(item)]
    failed = [item for item in review_results if not is_passed_result(item)]
    return passed, failed


def export_rows_for_mode(job: dict[str, Any], mode: str) -> tuple[list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]]]:
    paths = job.get('paths') or {}
    review_results = read_results_file(paths.get('review_results_file'))
    restore_results = read_results_file(paths.get('restore_results_file'))
    passed, failed = classify_review_results(review_results)

    if mode == 'passed':
        return passed, [], []
    if mode == 'failed':
        return failed, [], []
    if mode == 'all':
        return review_results, restore_results, []
    raise HTTPException(status_code=400, detail='导出类型不支持')


def normalize_result_row(item: dict[str, Any], source: str, job: dict[str, Any]) -> list[Any]:
    return [
        item.get('goods_id') or '',
        item.get('goods_name') or '',
        item.get('review_status') or item.get('status') or '',
        source,
        job.get('job_id') or '',
        job.get('created_at') or '',
        job.get('ended_at') or '',
    ]


def append_sheet(workbook: Workbook, title: str, rows: list[dict[str, Any]], source: str, job: dict[str, Any]) -> None:
    sheet = workbook.create_sheet(title=title[:31])
    headers = ['商品ID', '商品名称', '状态', '来源', '任务ID', '任务创建时间', '任务结束时间']
    sheet.append(headers)

    header_fill = PatternFill('solid', fgColor='E7EEF8')
    for cell in sheet[1]:
        cell.font = Font(bold=True, color='18212F')
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')

    for item in rows:
        sheet.append(normalize_result_row(item, source, job))

    if not rows:
        sheet.append(['', '', '暂无数据', source, job.get('job_id') or '', job.get('created_at') or '', job.get('ended_at') or ''])

    widths = [18, 42, 14, 14, 26, 22, 22]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical='top', wrap_text=True)
    sheet.freeze_panes = 'A2'


def build_export_workbook(job: dict[str, Any], mode: str) -> bytes:
    paths = job.get('paths') or {}
    review_results = read_results_file(paths.get('review_results_file'))
    restore_results = read_results_file(paths.get('restore_results_file'))
    passed, failed = classify_review_results(review_results)

    workbook = Workbook()
    workbook.remove(workbook.active)

    if mode == 'failed':
        append_sheet(workbook, '审核未通过', failed, '审核结果', job)
    elif mode == 'passed':
        append_sheet(workbook, '审核通过', passed, '审核结果', job)
    elif mode == 'all':
        append_sheet(workbook, '审核全部', review_results, '审核结果', job)
        append_sheet(workbook, '审核通过', passed, '审核结果', job)
        append_sheet(workbook, '审核未通过', failed, '审核结果', job)
        append_sheet(workbook, '恢复结果', restore_results, '恢复结果', job)
    else:
        raise HTTPException(status_code=400, detail='导出类型不支持')

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def clamp_int(value: Any, default: int, minimum: int = 1, maximum: int | None = None) -> int:
    try:
        number = int(value)
    except (TypeError, ValueError):
        number = default
    number = max(minimum, number)
    if maximum is not None:
        number = min(maximum, number)
    return number


def clean_filename(name: str, fallback: str) -> str:
    cleaned = ''.join(ch for ch in name if ch.isalnum() or ch in '._-（）()[] ')
    return cleaned.strip() or fallback


def has_upload(upload: UploadFile | None) -> bool:
    return bool(upload and upload.filename)


async def save_upload(upload: UploadFile, target: Path, allowed_suffixes: set[str]) -> None:
    suffix = Path(upload.filename or '').suffix.lower()
    if suffix not in allowed_suffixes:
        raise HTTPException(status_code=400, detail=f'文件格式不支持：{upload.filename}')
    target.parent.mkdir(parents=True, exist_ok=True)
    content = await upload.read()
    if not content:
        raise HTTPException(status_code=400, detail=f'文件为空：{upload.filename}')
    target.write_bytes(content)


def find_upload_entry(kind: str, upload_id: str) -> dict[str, Any] | None:
    state = load_state()
    for item in state.get('uploads', {}).get(kind, []):
        if item.get('id') == upload_id:
            return item
    return None


def ensure_upload_path(entry: dict[str, Any]) -> Path:
    path = Path(str(entry.get('path', ''))).resolve()
    allowed_roots = [UPLOADS_DIR.resolve(), APP_CONFIG.legacy_uploads_dir.resolve()]
    if not any(root == path or root in path.parents for root in allowed_roots):
        raise HTTPException(status_code=400, detail='历史文件路径不合法')
    if not path.exists() or not path.is_file():
        raise HTTPException(status_code=400, detail=f'历史文件不存在：{entry.get("original_name") or path.name}')
    return path


async def materialize_input_file(
    upload: UploadFile | None,
    selected_id: str,
    kind: str,
    target_prefix: str,
    fallback_name: str,
    input_dir: Path,
    allowed_suffixes: set[str],
) -> tuple[Path, dict[str, Any], dict[str, Any]]:
    selected_id = (selected_id or '').strip()

    if has_upload(upload):
        upload_id = uuid.uuid4().hex
        original_name = upload.filename or fallback_name
        stored_name = upload_id + '_' + clean_filename(original_name, fallback_name)
        stored_path = UPLOADS_DIR / kind / stored_name
        await save_upload(upload, stored_path, allowed_suffixes)
        target_path = input_dir / (target_prefix + clean_filename(original_name, fallback_name))
        target_path.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(stored_path, target_path)
        stat = stored_path.stat()
        entry = {
            'id': upload_id,
            'type': kind,
            'original_name': original_name,
            'path': str(stored_path),
            'size': stat.st_size,
            'created_at': now_iso(),
            'last_used_at': now_iso(),
        }
        return target_path, entry, {'action': 'add', 'kind': kind, 'entry': entry}

    if selected_id:
        entry = find_upload_entry(kind, selected_id)
        if not entry:
            raise HTTPException(status_code=400, detail='没有找到选择的历史文件')
        source_path = ensure_upload_path(entry)
        original_name = entry.get('original_name') or source_path.name
        target_path = input_dir / (target_prefix + clean_filename(original_name, fallback_name))
        target_path.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(source_path, target_path)
        return target_path, entry, {'action': 'touch', 'kind': kind, 'id': selected_id}

    raise HTTPException(status_code=400, detail=f'请上传或选择历史文件：{fallback_name}')


def append_job_log(job_id: str, line: str) -> None:
    line = line.rstrip('\r\n')
    with runtime_lock:
        runtime = job_runtimes.setdefault(job_id, {'logs': []})
        runtime.setdefault('logs', []).append(line)
        log_file = runtime.get('log_file')
    if log_file:
        Path(log_file).parent.mkdir(parents=True, exist_ok=True)
        with Path(log_file).open('a', encoding='utf-8') as file:
            file.write(line + '\n')


def update_job_record(job_id: str, **updates: Any) -> None:
    def mutate(state: dict[str, Any]) -> None:
        for job in state['jobs']:
            if job.get('job_id') == job_id:
                job.update(updates)
                return

    update_state(mutate)


def active_job_count() -> int:
    with runtime_lock:
        return sum(1 for item in job_runtimes.values() if item.get('status') in {'pending', 'running', 'paused', 'stopping'})


def comparable_path(path: str | Path) -> str:
    return os.path.normcase(os.path.abspath(os.path.expandvars(str(path).strip('"'))))


def command_uses_user_data_dir(cmdline: list[str], user_data_dir: Path) -> bool:
    target = comparable_path(user_data_dir)
    for index, arg in enumerate(cmdline):
        value: str | None = None
        if arg.startswith('--user-data-dir='):
            value = arg.split('=', 1)[1]
        elif arg == '--user-data-dir' and index + 1 < len(cmdline):
            value = cmdline[index + 1]
        if value and comparable_path(value) == target:
            return True
    return False


def chrome_profile_processes(user_data_dir: Path = DEFAULT_USER_DATA_DIR) -> list[psutil.Process]:
    processes: dict[int, psutil.Process] = {}
    for process in psutil.process_iter(['name', 'cmdline']):
        try:
            cmdline = process.info.get('cmdline') or []
            name = process.info.get('name') or ''
            if 'chrome' not in name.lower() and 'chromium' not in name.lower():
                continue
            if not command_uses_user_data_dir(cmdline, user_data_dir):
                continue
            processes[process.pid] = process
            for child in process.children(recursive=True):
                processes[child.pid] = child
        except psutil.Error:
            continue
    return list(processes.values())


def restart_automation_profile(log_callback) -> None:
    processes = chrome_profile_processes(DEFAULT_USER_DATA_DIR)
    if not processes:
        log_callback(f'启动前检查：未发现占用专用浏览器目录的 Chrome：{DEFAULT_USER_DATA_DIR}')
        return

    log_callback(f'启动前重启专用浏览器目录，将关闭 {len(processes)} 个 Chrome 进程：{DEFAULT_USER_DATA_DIR}')
    for process in processes:
        with contextlib.suppress(psutil.Error):
            process.resume()
    for process in processes:
        with contextlib.suppress(psutil.Error):
            process.terminate()

    _, alive = psutil.wait_procs(processes, timeout=6)
    if alive:
        log_callback(f'仍有 {len(alive)} 个 Chrome 进程未退出，执行强制关闭。')
    for process in alive:
        with contextlib.suppress(psutil.Error):
            process.kill()
    if alive:
        psutil.wait_procs(alive, timeout=3)


def find_free_port() -> int:
    for _ in range(20):
        with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as sock:
            sock.bind(('127.0.0.1', 0))
            port = int(sock.getsockname()[1])
        if port != 9222:
            return port
    raise RuntimeError('未能分配可用调试端口')


def process_tree(root_pid: int) -> list[psutil.Process]:
    try:
        root = psutil.Process(root_pid)
        return [*root.children(recursive=True), root]
    except psutil.Error:
        return []


def pause_process_tree(root_pid: int) -> None:
    for process in process_tree(root_pid):
        with contextlib.suppress(psutil.Error):
            process.suspend()


def resume_process_tree(root_pid: int) -> None:
    for process in process_tree(root_pid):
        with contextlib.suppress(psutil.Error):
            process.resume()


def terminate_process_tree(root_pid: int, timeout: int = 8) -> None:
    processes = process_tree(root_pid)
    for process in processes:
        with contextlib.suppress(psutil.Error):
            process.resume()
    for process in processes:
        with contextlib.suppress(psutil.Error):
            process.terminate()
    gone, alive = psutil.wait_procs(processes, timeout=timeout)
    for process in alive:
        with contextlib.suppress(psutil.Error):
            process.kill()


def get_running_process(job_id: str) -> subprocess.Popen:
    with runtime_lock:
        runtime = job_runtimes.get(job_id)
        if not runtime:
            raise HTTPException(status_code=404, detail='任务不在当前运行队列中')
        process = runtime.get('process')
        status = runtime.get('status')
    if process is None or process.poll() is not None:
        raise HTTPException(status_code=409, detail='任务进程未运行')
    if status not in {'running', 'paused', 'stopping'}:
        raise HTTPException(status_code=409, detail=f'当前状态不能操作：{status}')
    return process


def read_job_logs(job: dict[str, Any]) -> list[str]:
    log_file = Path(job.get('paths', {}).get('log_file', ''))
    if log_file.exists():
        return log_file.read_text(encoding='utf-8', errors='ignore').splitlines()
    with runtime_lock:
        runtime = job_runtimes.get(job.get('job_id', ''))
        if runtime:
            return list(runtime.get('logs', []))
    return []


def run_job_process(job_id: str, config_file: Path) -> None:
    command = [sys.executable, '-u', str(OPEN_SCRIPT), '--config', str(config_file)]
    env = os.environ.copy()
    env['PYTHONIOENCODING'] = 'utf-8'
    creationflags = getattr(subprocess, 'CREATE_NO_WINDOW', 0)
    config_payload = json.loads(config_file.read_text(encoding='utf-8'))

    with runtime_lock:
        runtime = job_runtimes[job_id]
        runtime['status'] = 'running'
        runtime['started_at'] = now_iso()
    update_job_record(job_id, status='running', started_at=now_iso())
    append_job_log(job_id, f'启动任务：{job_id}')
    append_job_log(job_id, f'浏览器目录：{DEFAULT_USER_DATA_DIR}')
    append_job_log(job_id, f'浏览器程序：{BROWSER_PATH.resolve() if BROWSER_PATH else "系统默认 Chrome"}')
    append_job_log(job_id, f'扩展目录：{EXTENSION_DIR}，存在：{EXTENSION_DIR.exists()}')
    append_job_log(job_id, f'调试端口：{config_payload.get("debug_port")}')
    restart_automation_profile(lambda message: append_job_log(job_id, message))

    try:
        process = subprocess.Popen(
            command,
            cwd=str(BASE_DIR),
            stdout=subprocess.PIPE,
            stderr=subprocess.STDOUT,
            text=True,
            encoding='utf-8',
            errors='replace',
            env=env,
            creationflags=creationflags,
        )
        with runtime_lock:
            job_runtimes[job_id]['process'] = process

        assert process.stdout is not None
        for line in process.stdout:
            append_job_log(job_id, line)
        returncode = process.wait()
        with runtime_lock:
            stop_requested = bool(job_runtimes.get(job_id, {}).get('stop_requested'))
        status = 'stopped' if stop_requested else ('completed' if returncode == 0 else 'failed')
        append_job_log(job_id, f'任务结束，退出码：{returncode}')
    except Exception as exc:
        returncode = -1
        status = 'failed'
        append_job_log(job_id, f'任务启动或执行失败：{exc}')

    ended_at = now_iso()
    with runtime_lock:
        runtime = job_runtimes[job_id]
        runtime['status'] = status
        runtime['ended_at'] = ended_at
        runtime['returncode'] = returncode
    update_job_record(job_id, status=status, ended_at=ended_at, returncode=returncode)


def append_login_log(line: str) -> None:
    line = line.rstrip('\r\n')
    with runtime_lock:
        login_runtime.setdefault('logs', []).append(line)


def run_login_process() -> None:
    debug_port = find_free_port()
    command = [
        sys.executable,
        '-u',
        str(LOGIN_SCRIPT),
        '--output',
        str(COOKIE_FILE),
        '--user-data-dir',
        str(DEFAULT_USER_DATA_DIR),
        '--debug-port',
        str(debug_port),
    ]
    if BROWSER_PATH:
        command.extend(['--browser-path', str(BROWSER_PATH)])
    env = os.environ.copy()
    env['PYTHONIOENCODING'] = 'utf-8'
    creationflags = getattr(subprocess, 'CREATE_NO_WINDOW', 0)

    with runtime_lock:
        login_runtime.update(
            {
                'status': 'running',
                'logs': [],
                'started_at': now_iso(),
                'ended_at': None,
                'returncode': None,
                'debug_port': debug_port,
            }
        )
    append_login_log('启动扫码登录流程')
    append_login_log(f'浏览器目录：{DEFAULT_USER_DATA_DIR}')
    append_login_log(f'浏览器程序：{BROWSER_PATH.resolve() if BROWSER_PATH else "系统默认 Chrome"}')
    append_login_log(f'扩展目录：{EXTENSION_DIR}，存在：{EXTENSION_DIR.exists()}')
    append_login_log(f'调试端口：{debug_port}')
    restart_automation_profile(append_login_log)

    try:
        process = subprocess.Popen(
            command,
            cwd=str(BASE_DIR),
            stdout=subprocess.PIPE,
            stderr=subprocess.STDOUT,
            text=True,
            encoding='utf-8',
            errors='replace',
            env=env,
            creationflags=creationflags,
        )
        assert process.stdout is not None
        for line in process.stdout:
            append_login_log(line)
        returncode = process.wait()
        status = 'completed' if returncode == 0 else 'failed'
        append_login_log(f'登录流程结束，退出码：{returncode}')
    except Exception as exc:
        returncode = -1
        status = 'failed'
        append_login_log(f'登录流程启动或执行失败：{exc}')

    with runtime_lock:
        login_runtime.update({'status': status, 'ended_at': now_iso(), 'returncode': returncode})


@app.get('/')
def index(request: Request):
    reconcile_stale_jobs()
    state = load_state()
    return templates.TemplateResponse(
        'index.html',
        {
            'request': request,
            'settings': state['last_settings'],
            'jobs': state['jobs'][:50],
            'uploads': state['uploads'],
            'cookie': cookie_status(),
            'extension': extension_status(),
        },
    )


@app.post('/login/start')
def start_login():
    env_errors = automation_environment_errors()
    if env_errors:
        raise HTTPException(status_code=400, detail='；'.join(env_errors))
    with runtime_lock:
        if login_runtime.get('status') == 'running':
            raise HTTPException(status_code=409, detail='登录流程正在运行')
        has_active_job = any(
            item.get('status') in {'pending', 'running', 'paused', 'stopping'}
            for item in job_runtimes.values()
        )
        if has_active_job:
            raise HTTPException(status_code=409, detail='自动化任务正在运行，不能同时刷新登录')
    thread = threading.Thread(target=run_login_process, daemon=True)
    thread.start()
    return {'ok': True}


@app.get('/api/login/status')
def get_login_status():
    with runtime_lock:
        runtime = dict(login_runtime)
        runtime['logs'] = list(login_runtime.get('logs', []))[-200:]
    runtime['cookie'] = cookie_status()
    runtime['extension'] = extension_status()
    return runtime


@app.get('/api/extension/status')
def get_extension_status():
    return extension_status()


@app.get('/api/uploads')
def list_uploads():
    state = load_state()
    return {'uploads': state.get('uploads', {'excel': [], 'images': []})}


@app.post('/jobs')
async def create_job(
    excel_file: UploadFile | None = File(None),
    image_file_1: UploadFile | None = File(None),
    image_file_2: UploadFile | None = File(None),
    existing_excel_id: str = Form(''),
    existing_image_1_id: str = Form(''),
    existing_image_2_id: str = Form(''),
    text_to_find: str = Form(...),
    text_to_replace: str = Form(...),
    flash_sale_amount: str = Form('10'),
    max_batch_size: int = Form(50),
    concurrency: int = Form(1),
    review_timeout: int = Form(80),
    default_timeout: int = Form(15),
    poll_interval: int = Form(1),
    plugin_ready_retry_times: int = Form(10),
    plugin_ready_retry_interval: int = Form(2),
    search_result_retry_times: int = Form(5),
    search_result_retry_interval: int = Form(3),
    bulk_action_retry_times: int = Form(5),
    bulk_action_retry_interval: int = Form(1),
    select_all_retry_times: int = Form(5),
    select_all_retry_interval: int = Form(1),
    restore_dialog_retry_times: int = Form(5),
    restore_dialog_retry_interval: int = Form(3),
    max_parallel_jobs: int = Form(1),
    browser_profile_mode: str = Form('shared'),
):
    env_errors = automation_environment_errors()
    if env_errors:
        raise HTTPException(status_code=400, detail='；'.join(env_errors))
    if not COOKIE_FILE.exists():
        if APP_CONFIG.legacy_cookie_file.exists():
            raise HTTPException(
                status_code=400,
                detail=f'缺少 {COOKIE_FILE}。检测到旧 Cookie：{APP_CONFIG.legacy_cookie_file}，请刷新登录或手动迁移。',
            )
        raise HTTPException(status_code=400, detail=f'缺少 {COOKIE_FILE}，请先刷新登录')
    if not text_to_find.strip() or not text_to_replace.strip():
        raise HTTPException(status_code=400, detail='原文和新文不能为空')

    max_parallel_jobs = 1
    browser_profile_mode = 'shared'
    with runtime_lock:
        if login_runtime.get('status') == 'running':
            raise HTTPException(status_code=409, detail='登录流程正在运行，不能同时启动任务')
    if active_job_count() > 0:
        raise HTTPException(status_code=409, detail='专用浏览器目录一次只能运行一个任务')
    if active_job_count() >= max_parallel_jobs:
        raise HTTPException(status_code=409, detail=f'当前已达到完整任务并行上限：{max_parallel_jobs}')

    job_id = datetime.now().strftime('%Y%m%d-%H%M%S') + '-' + uuid.uuid4().hex[:8]
    job_dir = RUNS_DIR / job_id
    input_dir = job_dir / 'inputs'
    review_results_file = job_dir / 'review_results.json'
    restore_results_file = job_dir / 'restore_results.json'
    log_file = job_dir / 'run.log'
    config_file = job_dir / 'config.json'
    debug_port = find_free_port()

    upload_updates: list[dict[str, Any]] = []
    excel_path, excel_entry, upload_update = await materialize_input_file(
        excel_file,
        existing_excel_id,
        'excel',
        'excel_',
        'goods.xlsx',
        input_dir,
        {'.xlsx'},
    )
    upload_updates.append(upload_update)
    image_path_1, image_entry_1, upload_update = await materialize_input_file(
        image_file_1,
        existing_image_1_id,
        'images',
        'image_1_',
        'image_1.jpg',
        input_dir,
        {'.jpg', '.jpeg', '.png', '.webp'},
    )
    upload_updates.append(upload_update)
    image_path_2, image_entry_2, upload_update = await materialize_input_file(
        image_file_2,
        existing_image_2_id,
        'images',
        'image_2_',
        'image_2.jpg',
        input_dir,
        {'.jpg', '.jpeg', '.png', '.webp'},
    )
    upload_updates.append(upload_update)

    settings = {
        'max_batch_size': clamp_int(max_batch_size, 50),
        'concurrency': clamp_int(concurrency, 1),
        'review_timeout': clamp_int(review_timeout, 80),
        'default_timeout': clamp_int(default_timeout, 15),
        'poll_interval': clamp_int(poll_interval, 1),
        'plugin_ready_retry_times': clamp_int(plugin_ready_retry_times, 10),
        'plugin_ready_retry_interval': clamp_int(plugin_ready_retry_interval, 2),
        'search_result_retry_times': clamp_int(search_result_retry_times, 5),
        'search_result_retry_interval': clamp_int(search_result_retry_interval, 3),
        'bulk_action_retry_times': clamp_int(bulk_action_retry_times, 5),
        'bulk_action_retry_interval': clamp_int(bulk_action_retry_interval, 1),
        'select_all_retry_times': clamp_int(select_all_retry_times, 5),
        'select_all_retry_interval': clamp_int(select_all_retry_interval, 1),
        'restore_dialog_retry_times': clamp_int(restore_dialog_retry_times, 5),
        'restore_dialog_retry_interval': clamp_int(restore_dialog_retry_interval, 3),
        'flash_sale_amount': flash_sale_amount.strip() or '10',
        'text_to_find': text_to_find.strip(),
        'text_to_replace': text_to_replace.strip(),
        'max_parallel_jobs': max_parallel_jobs,
        'browser_profile_mode': browser_profile_mode,
    }
    user_data_dir = DEFAULT_USER_DATA_DIR
    config = {
        **settings,
        'cookie_file': str(COOKIE_FILE),
        'excel_file': str(excel_path),
        'user_data_dir': str(user_data_dir),
        'debug_port': debug_port,
        'browser_path': str(BROWSER_PATH) if BROWSER_PATH else '',
        'review_results_file': str(review_results_file),
        'restore_results_file': str(restore_results_file),
        'image_path_1': str(image_path_1),
        'image_path_2': str(image_path_2),
    }
    config_file.write_text(json.dumps(config, ensure_ascii=False, indent=2), encoding='utf-8')

    job_record = {
        'job_id': job_id,
        'status': 'pending',
        'created_at': now_iso(),
        'started_at': None,
        'ended_at': None,
        'returncode': None,
        'excel_file': excel_entry.get('original_name'),
        'image_file_1': image_entry_1.get('original_name'),
        'image_file_2': image_entry_2.get('original_name'),
        'summary': {
            'max_batch_size': settings['max_batch_size'],
            'concurrency': settings['concurrency'],
            'review_timeout': settings['review_timeout'],
            'flash_sale_amount': settings['flash_sale_amount'],
            'browser_profile_mode': settings['browser_profile_mode'],
            'debug_port': debug_port,
            'browser_path': str(BROWSER_PATH) if BROWSER_PATH else '',
        },
        'paths': {
            'job_dir': str(job_dir),
            'review_results_file': str(review_results_file),
            'restore_results_file': str(restore_results_file),
            'log_file': str(log_file),
        },
    }

    def mutate(state: dict[str, Any]) -> None:
        state['last_settings'] = settings
        uploads = state.setdefault('uploads', {'excel': [], 'images': []})
        uploads.setdefault('excel', [])
        uploads.setdefault('images', [])
        for update in upload_updates:
            items = uploads[update['kind']]
            if update['action'] == 'add':
                items.insert(0, update['entry'])
            elif update['action'] == 'touch':
                for item in items:
                    if item.get('id') == update['id']:
                        item['last_used_at'] = now_iso()
                        break
            uploads[update['kind']] = items[:50]
        state['jobs'].insert(0, job_record)
        state['jobs'] = state['jobs'][:100]

    update_state(mutate)
    with runtime_lock:
        job_runtimes[job_id] = {
            'status': 'pending',
            'logs': [],
            'log_file': str(log_file),
            'created_at': job_record['created_at'],
        }

    thread = threading.Thread(target=run_job_process, args=(job_id, config_file), daemon=True)
    thread.start()
    return JSONResponse({'ok': True, 'job_id': job_id})


@app.get('/api/jobs')
def list_jobs():
    reconcile_stale_jobs()
    state = load_state()
    return {'jobs': [enrich_job(job) for job in state.get('jobs', [])]}


@app.get('/api/jobs/{job_id}')
def get_job(job_id: str):
    job = find_job(job_id)
    logs = read_job_logs(job)
    return {'job': enrich_job(job), 'logs': logs, 'log_count': len(logs)}


@app.get('/api/jobs/{job_id}/export')
def export_job_results(job_id: str, mode: str = 'all'):
    mode = (mode or 'all').strip().lower()
    labels = {
        'failed': '审核未通过',
        'passed': '审核通过',
        'all': '全部结果',
    }
    if mode not in labels:
        raise HTTPException(status_code=400, detail='导出类型不支持')

    job = find_job(job_id)
    content = build_export_workbook(job, mode)
    filename = f'{job_id}_{labels[mode]}.xlsx'
    encoded = quote(filename)
    headers = {
        'Content-Disposition': f"attachment; filename*=UTF-8''{encoded}",
    }
    return StreamingResponse(
        BytesIO(content),
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers=headers,
    )


@app.post('/api/jobs/{job_id}/pause')
def pause_job(job_id: str):
    process = get_running_process(job_id)
    with runtime_lock:
        current_status = job_runtimes[job_id].get('status')
    if current_status == 'paused':
        return {'ok': True, 'status': 'paused'}
    if current_status != 'running':
        raise HTTPException(status_code=409, detail=f'当前状态不能暂停：{current_status}')
    pause_process_tree(process.pid)
    with runtime_lock:
        job_runtimes[job_id]['status'] = 'paused'
        job_runtimes[job_id]['paused_at'] = now_iso()
    update_job_record(job_id, status='paused', paused_at=now_iso())
    append_job_log(job_id, '任务已暂停')
    return {'ok': True, 'status': 'paused'}


@app.post('/api/jobs/{job_id}/resume')
def resume_job(job_id: str):
    process = get_running_process(job_id)
    with runtime_lock:
        current_status = job_runtimes[job_id].get('status')
    if current_status == 'running':
        return {'ok': True, 'status': 'running'}
    if current_status != 'paused':
        raise HTTPException(status_code=409, detail=f'当前状态不能继续：{current_status}')
    resume_process_tree(process.pid)
    with runtime_lock:
        job_runtimes[job_id]['status'] = 'running'
        job_runtimes[job_id]['resumed_at'] = now_iso()
    update_job_record(job_id, status='running', resumed_at=now_iso())
    append_job_log(job_id, '任务已继续')
    return {'ok': True, 'status': 'running'}


@app.post('/api/jobs/{job_id}/stop')
def stop_job(job_id: str):
    process = get_running_process(job_id)
    with runtime_lock:
        job_runtimes[job_id]['status'] = 'stopping'
        job_runtimes[job_id]['stop_requested'] = True
    update_job_record(job_id, status='stopping')
    append_job_log(job_id, '正在停止任务...')
    terminate_process_tree(process.pid)
    return {'ok': True, 'status': 'stopping'}


@app.get('/api/jobs/{job_id}/events')
async def job_events(job_id: str, from_index: int = 0):
    async def stream():
        index = max(0, from_index)
        while True:
            with runtime_lock:
                runtime = job_runtimes.get(job_id)
                logs = list(runtime.get('logs', [])) if runtime else []
                status = runtime.get('status') if runtime else None
            for line in logs[index:]:
                yield 'data: ' + json.dumps({'line': line}, ensure_ascii=False) + '\n\n'
            index = len(logs)
            if status in {'completed', 'failed', 'stopped'}:
                yield 'event: done\ndata: ' + json.dumps({'status': status}, ensure_ascii=False) + '\n\n'
                break
            await asyncio.sleep(1)

    return StreamingResponse(stream(), media_type='text/event-stream')


if __name__ == '__main__':
    import uvicorn

    APP_CONFIG.ensure_data_dirs()
    uvicorn.run(app, host='127.0.0.1', port=8000, reload=False)

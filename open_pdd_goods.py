import argparse
import contextlib
import json
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass
from pathlib import Path
from typing import Callable
from urllib.parse import urlparse

from openpyxl import load_workbook
from DrissionPage import ChromiumOptions, ChromiumPage

from app_config import APP_CONFIG

BASE_DIR = APP_CONFIG.base_dir
EXTENSION_DIR = APP_CONFIG.extension_dir
USER_DATA_DIR = APP_CONFIG.user_data_dir
DEBUG_PORT: int | None = None
BROWSER_PATH: Path | None = APP_CONFIG.browser_path
COOKIE_FILE = APP_CONFIG.cookie_file
HOME_URL = 'https://mms.pinduoduo.com/'
TARGET_URL = 'https://mms.pinduoduo.com/goods/goods_list?msfrom=mms_sidenav'
EXTENSIONS_URL = 'chrome://extensions/'
MANUAL_EXTENSION_WAIT_SECONDS = 30
EXCEL_GLOB = '*.xlsx'
MAX_BATCH_SIZE = 50
REVIEW_TIMEOUT = 80
DEFAULT_TIMEOUT = 15
POLL_INTERVAL = 1
PLUGIN_READY_RETRY_TIMES = 10
PLUGIN_READY_RETRY_INTERVAL = 2
SEARCH_RESULT_RETRY_TIMES = 5
SEARCH_RESULT_RETRY_INTERVAL = 3
BULK_ACTION_RETRY_TIMES = 5
BULK_ACTION_RETRY_INTERVAL = 1
SELECT_ALL_RETRY_TIMES = 5
SELECT_ALL_RETRY_INTERVAL = 1
RESTORE_DIALOG_RETRY_TIMES = 5
RESTORE_DIALOG_RETRY_INTERVAL = 3
REVIEW_RESULTS_FILE = Path(__file__).with_name('pdd_review_results.json')
FLASH_SALE_AMOUNT = '10'
TEXT_TO_FIND = None
TEXT_TO_REPLACE = None
IMAGE_PATHS = (None, None)


@dataclass(slots=True)
class AutomationConfig:
    cookie_file: Path = COOKIE_FILE
    excel_file: Path | None = None
    user_data_dir: Path = USER_DATA_DIR
    debug_port: int | None = None
    browser_path: Path | None = None
    review_results_file: Path = REVIEW_RESULTS_FILE
    restore_results_file: Path = Path(__file__).with_name('pdd_restore_results.json')
    max_batch_size: int = MAX_BATCH_SIZE
    concurrency: int = 1
    review_timeout: int = REVIEW_TIMEOUT
    default_timeout: int = DEFAULT_TIMEOUT
    poll_interval: int = POLL_INTERVAL
    plugin_ready_retry_times: int = PLUGIN_READY_RETRY_TIMES
    plugin_ready_retry_interval: int = PLUGIN_READY_RETRY_INTERVAL
    search_result_retry_times: int = SEARCH_RESULT_RETRY_TIMES
    search_result_retry_interval: int = SEARCH_RESULT_RETRY_INTERVAL
    bulk_action_retry_times: int = BULK_ACTION_RETRY_TIMES
    bulk_action_retry_interval: int = BULK_ACTION_RETRY_INTERVAL
    select_all_retry_times: int = SELECT_ALL_RETRY_TIMES
    select_all_retry_interval: int = SELECT_ALL_RETRY_INTERVAL
    restore_dialog_retry_times: int = RESTORE_DIALOG_RETRY_TIMES
    restore_dialog_retry_interval: int = RESTORE_DIALOG_RETRY_INTERVAL
    flash_sale_amount: str = FLASH_SALE_AMOUNT
    text_to_find: str | None = None
    text_to_replace: str | None = None
    image_paths: tuple[Path | None, Path | None] = (None, None)

    @classmethod
    def from_dict(cls, payload: dict) -> 'AutomationConfig':
        def as_path(key: str, default: Path | None = None) -> Path | None:
            value = payload.get(key)
            if value in (None, ''):
                return default
            return Path(str(value))

        def as_int(key: str, default: int, minimum: int = 1) -> int:
            value = payload.get(key, default)
            try:
                number = int(value)
            except (TypeError, ValueError):
                number = default
            return max(minimum, number)

        def as_optional_int(key: str, minimum: int = 1) -> int | None:
            value = payload.get(key)
            if value in (None, ''):
                return None
            try:
                number = int(value)
            except (TypeError, ValueError):
                return None
            return number if number >= minimum else None

        image_1 = as_path('image_path_1')
        image_2 = as_path('image_path_2')
        image_paths_value = payload.get('image_paths')
        if isinstance(image_paths_value, list) and len(image_paths_value) >= 2:
            image_1 = Path(str(image_paths_value[0])) if image_paths_value[0] else image_1
            image_2 = Path(str(image_paths_value[1])) if image_paths_value[1] else image_2

        default_restore_file = Path(__file__).with_name('pdd_restore_results.json')
        return cls(
            cookie_file=as_path('cookie_file', COOKIE_FILE) or COOKIE_FILE,
            excel_file=as_path('excel_file'),
            user_data_dir=as_path('user_data_dir', USER_DATA_DIR) or USER_DATA_DIR,
            debug_port=as_optional_int('debug_port'),
            browser_path=as_path('browser_path'),
            review_results_file=as_path('review_results_file', REVIEW_RESULTS_FILE) or REVIEW_RESULTS_FILE,
            restore_results_file=as_path('restore_results_file', default_restore_file) or default_restore_file,
            max_batch_size=as_int('max_batch_size', MAX_BATCH_SIZE),
            concurrency=as_int('concurrency', 1),
            review_timeout=as_int('review_timeout', REVIEW_TIMEOUT),
            default_timeout=as_int('default_timeout', DEFAULT_TIMEOUT),
            poll_interval=as_int('poll_interval', POLL_INTERVAL),
            plugin_ready_retry_times=as_int('plugin_ready_retry_times', PLUGIN_READY_RETRY_TIMES),
            plugin_ready_retry_interval=as_int('plugin_ready_retry_interval', PLUGIN_READY_RETRY_INTERVAL),
            search_result_retry_times=as_int('search_result_retry_times', SEARCH_RESULT_RETRY_TIMES),
            search_result_retry_interval=as_int('search_result_retry_interval', SEARCH_RESULT_RETRY_INTERVAL),
            bulk_action_retry_times=as_int('bulk_action_retry_times', BULK_ACTION_RETRY_TIMES),
            bulk_action_retry_interval=as_int('bulk_action_retry_interval', BULK_ACTION_RETRY_INTERVAL),
            select_all_retry_times=as_int('select_all_retry_times', SELECT_ALL_RETRY_TIMES),
            select_all_retry_interval=as_int('select_all_retry_interval', SELECT_ALL_RETRY_INTERVAL),
            restore_dialog_retry_times=as_int('restore_dialog_retry_times', RESTORE_DIALOG_RETRY_TIMES),
            restore_dialog_retry_interval=as_int('restore_dialog_retry_interval', RESTORE_DIALOG_RETRY_INTERVAL),
            flash_sale_amount=str(payload.get('flash_sale_amount') or FLASH_SALE_AMOUNT),
            text_to_find=payload.get('text_to_find') or None,
            text_to_replace=payload.get('text_to_replace') or None,
            image_paths=(image_1, image_2),
        )


class CallbackWriter:
    def __init__(self, callback: Callable[[str], None]):
        self.callback = callback
        self.buffer = ''

    def write(self, text: str) -> int:
        if not text:
            return 0
        self.buffer += text
        while '\n' in self.buffer:
            line, self.buffer = self.buffer.split('\n', 1)
            self.callback(line)
        return len(text)

    def flush(self) -> None:
        if self.buffer:
            self.callback(self.buffer)
            self.buffer = ''


def runtime_timeout(timeout: int | None) -> int:
    return DEFAULT_TIMEOUT if timeout is None else timeout


def apply_runtime_config(config: AutomationConfig) -> None:
    global USER_DATA_DIR, DEBUG_PORT, BROWSER_PATH, MAX_BATCH_SIZE, REVIEW_TIMEOUT, DEFAULT_TIMEOUT, POLL_INTERVAL
    global PLUGIN_READY_RETRY_TIMES, PLUGIN_READY_RETRY_INTERVAL
    global SEARCH_RESULT_RETRY_TIMES, SEARCH_RESULT_RETRY_INTERVAL
    global BULK_ACTION_RETRY_TIMES, BULK_ACTION_RETRY_INTERVAL
    global SELECT_ALL_RETRY_TIMES, SELECT_ALL_RETRY_INTERVAL
    global RESTORE_DIALOG_RETRY_TIMES, RESTORE_DIALOG_RETRY_INTERVAL
    global FLASH_SALE_AMOUNT, TEXT_TO_FIND, TEXT_TO_REPLACE, IMAGE_PATHS

    USER_DATA_DIR = config.user_data_dir
    DEBUG_PORT = config.debug_port
    BROWSER_PATH = config.browser_path
    MAX_BATCH_SIZE = config.max_batch_size
    REVIEW_TIMEOUT = config.review_timeout
    DEFAULT_TIMEOUT = config.default_timeout
    POLL_INTERVAL = config.poll_interval
    PLUGIN_READY_RETRY_TIMES = config.plugin_ready_retry_times
    PLUGIN_READY_RETRY_INTERVAL = config.plugin_ready_retry_interval
    SEARCH_RESULT_RETRY_TIMES = config.search_result_retry_times
    SEARCH_RESULT_RETRY_INTERVAL = config.search_result_retry_interval
    BULK_ACTION_RETRY_TIMES = config.bulk_action_retry_times
    BULK_ACTION_RETRY_INTERVAL = config.bulk_action_retry_interval
    SELECT_ALL_RETRY_TIMES = config.select_all_retry_times
    SELECT_ALL_RETRY_INTERVAL = config.select_all_retry_interval
    RESTORE_DIALOG_RETRY_TIMES = config.restore_dialog_retry_times
    RESTORE_DIALOG_RETRY_INTERVAL = config.restore_dialog_retry_interval
    FLASH_SALE_AMOUNT = config.flash_sale_amount
    TEXT_TO_FIND = config.text_to_find
    TEXT_TO_REPLACE = config.text_to_replace
    IMAGE_PATHS = tuple(str(path) if path else None for path in config.image_paths)

PLUGIN_TRIGGER_XPATH = 'xpath://*[@id="pdd-ext-root"]/div/div/div/span'
PLUGIN_ROOT_XPATH = 'xpath://*[@id="pdd-ext-root"]'
GOODS_MANAGEMENT_XPATH = 'xpath://*[@id="pdd-ext-root"]/div[2]/div[2]/div/div/div/div/div[2]/nav/div[2]/div[2]/span[2]'
SEARCH_INPUT_XPATH = 'xpath://*[@id="pdd-ext-root"]/div[2]/div[2]/div/div/div/div/div[2]/div/div/div/div[1]/div[1]/span/input'
TABLE_ROW_XPATH = 'xpath://*[@id="pdd-ext-root"]/div[2]/div[2]/div/div/div/div/div[2]/div/div/div/div[2]/div/div/div/div/div/div[2]/table/tbody/tr'
TABLE_CHECKBOX_CONTAINER_XPATH = 'xpath://*[@id="pdd-ext-root"]/div[2]/div[2]/div/div/div/div/div[2]/div/div/div/div[2]/div/div/div/div/div/div[1]'
TABLE_CHECKBOX_LABEL_XPATH = 'xpath://*[@id="pdd-ext-root"]/div[2]/div[2]/div/div/div/div/div[2]/div/div/div/div[2]/div/div/div/div/div/div[1]/table/thead/tr/th[1]/label'
TABLE_CHECKBOX_INPUT_XPATH = 'xpath://*[@id="pdd-ext-root"]/div[2]/div[2]/div/div/div/div/div[2]/div/div/div/div[2]/div/div/div/div/div/div[1]/table/thead/tr/th[1]/label/span/input'
BULK_ACTION_BUTTON_XPATH = 'xpath://button[.//span[contains(normalize-space(text()), "批量操作")] or contains(normalize-space(text()), "批量操作")]'
END_FLASH_SALE_MENU_XPATH = 'xpath://*[@id="pdd-ext-root"]/div[3]/ul/li[3]'
END_FLASH_SALE_CONFIRM_XPATH = 'xpath://*[@id="pdd-ext-root"]/div[4]/div[2]/div/div/div[2]/div/div[2]/button[2]'
BULK_EDIT_MENU_XPATH = 'xpath://*[@id="pdd-ext-root"]/div[3]/ul/li[6]'
ORIGINAL_TEXT_INPUT_XPATH = 'xpath://*[@id="pdd-ext-root"]/div[4]/div[2]/div/div/div/div/div[2]/div[2]/div[1]/div[2]/div[2]/input[1]'
REPLACEMENT_TEXT_INPUT_XPATH = 'xpath://*[@id="pdd-ext-root"]/div[4]/div[2]/div/div/div/div/div[2]/div[2]/div[1]/div[2]/div[2]/input[2]'
TEXT_REPLACE_BUTTON_XPATH = 'xpath://*[@id="pdd-ext-root"]/div[4]/div[2]/div/div/div/div/div[2]/div[2]/div[1]/div[2]/div[2]/button[1]'
IMAGE_UPLOAD_1_XPATH = 'xpath://*[@id="pdd-ext-root"]/div[4]/div[2]/div/div/div/div/div[2]/div[2]/div[1]/div[3]/div[2]/span[1]/div[2]'
IMAGE_UPLOAD_2_XPATH = 'xpath://*[@id="pdd-ext-root"]/div[4]/div[2]/div/div/div/div/div[2]/div[2]/div[1]/div[3]/div[2]/span[2]/div[2]'
EXECUTE_IMAGE_REPLACE_XPATH = 'xpath://*[@id="pdd-ext-root"]/div[4]/div[2]/div/div/div/div/div[2]/div[2]/div[1]/div[3]/div[2]/button/span'
SUBMIT_BUTTON_XPATH = 'xpath://*[@id="pdd-ext-root"]/div[4]/div[2]/div/div/div/div/div[3]/div/div[3]/button'
REVIEW_DIALOG_XPATH = 'xpath://div[@role="dialog" and contains(@class, "ant-modal")][.//span[contains(text(), "提交进度")]]'
REVIEW_CARD_XPATH = 'xpath:.//div[contains(@class, "sp-card")]'
REVIEW_NAME_XPATH = 'xpath:.//div[contains(@class, "sp-name")]'
REVIEW_ID_XPATH = 'xpath:.//div[contains(@class, "sp-id")]'
REVIEW_STATUS_XPATH = 'xpath:.//span[contains(@class, "sp-tag")]'
REVIEW_CLOSE_BUTTON_XPATH = 'xpath://*[@id="pdd-ext-root"]/div[4]/div[2]/div/div/div/div/div[3]/div/div[1]/button/span'
FLASH_SALE_BUTTON_XPATH = 'xpath://*[@id="pdd-ext-root"]/div[2]/div[2]/div/div/div/div/div[2]/div/div/div/div[1]/div[1]/button[7]'
FLASH_SALE_AMOUNT_INPUT_XPATH = 'xpath://*[@id="pdd-ext-root"]/div[3]/div[2]/div/div/div/div/div[2]/div[1]/div[2]/div[2]/div/input'
FLASH_SALE_APPLY_BUTTON_XPATH = 'xpath://*[@id="pdd-ext-root"]/div[3]/div[2]/div/div/div/div/div[2]/div[1]/div[2]/button'
FLASH_SALE_CREATE_BUTTON_XPATH = 'xpath://*[@id="pdd-ext-root"]/div[3]/div[2]/div/div/div/div/div[3]/div[2]/div[2]/button'
RESTORE_SELECTION_BUTTON_XPATH = 'xpath://*[@id="pdd-ext-root"]/div[2]/div[2]/div/div/div/div/div[2]/div/div/div/div[1]/div[1]/button[4]'
RESTORE_DIALOG_XPATH = 'xpath://div[@role="dialog" and contains(@class, "ant-modal")][.//span[contains(text(), "恢复图片备份")]]'


def load_json_file(file_path: Path) -> dict:
    if not file_path.exists():
        return {}
    try:
        return json.loads(file_path.read_text(encoding='utf-8', errors='ignore'))
    except json.JSONDecodeError:
        return {}


def get_registered_extension_paths() -> list[str]:
    preference_files = [
        USER_DATA_DIR / 'Default' / 'Preferences',
        USER_DATA_DIR / 'Default' / 'Secure Preferences',
    ]
    registered_paths: list[str] = []

    for preference_file in preference_files:
        data = load_json_file(preference_file)
        settings = (((data.get('extensions') or {}).get('settings')) or {})
        for config in settings.values():
            path_value = config.get('path')
            if isinstance(path_value, str) and path_value not in registered_paths:
                registered_paths.append(path_value)

    return registered_paths


def validate_extension_registration() -> None:
    expected_path = str(EXTENSION_DIR.resolve())
    registered_paths = get_registered_extension_paths()
    if not registered_paths:
        return
    if expected_path in registered_paths:
        return

    registered_lines = '\n'.join(registered_paths)
    print(
        '[警告] 固定用户目录中的扩展注册路径与当前扩展目录不一致，插件可能无法注入页面。\n'
        f'  当前扩展目录：{expected_path}\n'
        f'  已注册路径：{registered_lines}\n'
        '  如果插件入口找不到，请在浏览器扩展管理页删除旧的富多扩展，再从当前目录重新加载。'
    )


def log_startup_diagnostics() -> None:
    registered_paths = get_registered_extension_paths()
    registered_text = ', '.join(registered_paths) if registered_paths else '-'
    browser_text = str(BROWSER_PATH.resolve()) if BROWSER_PATH else '系统默认 Chrome'
    print(f'浏览器目录：{USER_DATA_DIR.resolve()}')
    print(f'浏览器程序：{browser_text}')
    print(f'扩展目录：{EXTENSION_DIR.resolve()}，存在：{EXTENSION_DIR.exists()}')
    print(f'调试端口：{DEBUG_PORT or "DrissionPage 默认"}')
    print(
        'Chrome 扩展启动参数：'
        f'--load-extension={EXTENSION_DIR.resolve()}，'
        f'--disable-extensions-except={EXTENSION_DIR.resolve()}，'
        '--remote-allow-origins=*'
    )
    print(f'profile 已注册扩展路径：{registered_text}')


def build_page() -> ChromiumPage:
    config_errors = list(APP_CONFIG.errors)
    if config_errors and BROWSER_PATH == APP_CONFIG.browser_path and EXTENSION_DIR == APP_CONFIG.extension_dir:
        raise RuntimeError('；'.join(config_errors))
    if not EXTENSION_DIR.exists():
        raise FileNotFoundError(f'未找到扩展目录：{EXTENSION_DIR}')
    if BROWSER_PATH and not BROWSER_PATH.exists():
        raise FileNotFoundError(f'未找到浏览器程序：{BROWSER_PATH}')

    validate_extension_registration()

    options = ChromiumOptions()
    if BROWSER_PATH:
        options.set_browser_path(str(BROWSER_PATH))
    options.set_user_data_path(str(USER_DATA_DIR))
    if DEBUG_PORT:
        options.set_local_port(DEBUG_PORT)
    options.set_argument('--remote-allow-origins', '*')
    options.add_extension(str(EXTENSION_DIR))
    options.headless(False)
    return ChromiumPage(addr_or_opts=options)


def wait_for_manual_extension_install() -> None:
    # print(f'正在打开扩展管理页：{EXTENSIONS_URL}')
    # page.get(EXTENSIONS_URL, timeout=DEFAULT_TIMEOUT)
    # print(f'请在 {MANUAL_EXTENSION_WAIT_SECONDS} 秒内手动安装扩展目录：{EXTENSION_DIR}')
    # print('安装完成后脚本会自动继续。')
    # time.sleep(MANUAL_EXTENSION_WAIT_SECONDS)
    return None


def load_cookies(cookie_file: Path) -> list[dict]:
    payload = json.loads(cookie_file.read_text(encoding='utf-8'))
    cookies = payload.get('cookies')
    if not isinstance(cookies, list) or not cookies:
        raise ValueError(f'未在 {cookie_file} 中找到可用 cookies')
    return cookies


def resolve_excel_file() -> Path:
    candidates = [path for path in BASE_DIR.glob(EXCEL_GLOB) if not path.name.startswith('~$')]
    if not candidates:
        raise FileNotFoundError('当前目录未找到可用的 Excel 文件')
    return candidates[0]


def normalize_goods_id(value: object) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    if text.endswith('.0'):
        text = text[:-2]
    digits = ''.join(ch for ch in text if ch.isdigit())
    return digits or None


def load_goods_ids_from_excel(excel_file: Path) -> list[str]:
    workbook = load_workbook(excel_file, read_only=True, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    goods_ids: list[str] = []
    seen: set[str] = set()

    try:
        for row in worksheet.iter_rows(values_only=True):
            for cell in row:
                goods_id = normalize_goods_id(cell)
                if goods_id and goods_id not in seen:
                    seen.add(goods_id)
                    goods_ids.append(goods_id)
    finally:
        workbook.close()

    if not goods_ids:
        raise ValueError(f'未从 {excel_file} 读取到商品 ID')
    return goods_ids


def chunk_goods_ids(goods_ids: list[str], batch_size: int = MAX_BATCH_SIZE) -> list[list[str]]:
    return [goods_ids[index:index + batch_size] for index in range(0, len(goods_ids), batch_size)]


def normalized_path(url: str) -> str:
    parsed_url = urlparse(url)
    path = parsed_url.path.rstrip('/')
    return path or '/'


def ensure_element(page: ChromiumPage, locator: str, timeout: int | None = None):
    timeout = runtime_timeout(timeout)
    element = page.ele(locator, timeout=timeout)
    if not element:
        raise RuntimeError(f'未找到元素：{locator}')
    return element


def wait_for_clickable(page: ChromiumPage, locator: str, timeout: int | None = None):
    timeout = runtime_timeout(timeout)
    element = ensure_element(page, locator, timeout=timeout)
    element.wait.clickable(timeout=timeout)
    return element


def click_element(page: ChromiumPage, locator: str, timeout: int | None = None) -> None:
    element = wait_for_clickable(page, locator, timeout=timeout)
    element.click()


def input_text(page: ChromiumPage, locator: str, value: str, timeout: int | None = None, clear: bool = True) -> None:
    element = wait_for_clickable(page, locator, timeout=timeout)
    element.input(value, clear=clear)




def click_element_by_text_contains(page: ChromiumPage, tag: str, text: str, timeout: int | None = None) -> None:
    locator = f'xpath://{tag}[contains(normalize-space(.), "{text}")]'
    click_element(page, locator, timeout=timeout)


def wait_for_element(page: ChromiumPage, locator: str, timeout: int | None = None) -> bool:
    timeout = runtime_timeout(timeout)
    element = page.ele(locator, timeout=timeout)
    return bool(element)


def log_extension_injection_failure(page: ChromiumPage) -> None:
    registered_paths = get_registered_extension_paths()
    registered_text = ', '.join(registered_paths) if registered_paths else '-'
    browser_text = str(BROWSER_PATH.resolve()) if BROWSER_PATH else '系统默认 Chrome'
    print('[错误] 扩展未注入：页面中没有找到 #pdd-ext-root。')
    print(f'当前页面：{page.url or "-"}')
    print(f'浏览器目录：{USER_DATA_DIR.resolve()}')
    print(f'浏览器程序：{browser_text}')
    print(f'扩展目录：{EXTENSION_DIR.resolve()}，存在：{EXTENSION_DIR.exists()}')
    print(f'调试端口：{DEBUG_PORT or "DrissionPage 默认"}')
    print(
        '期望 Chrome 启动参数：'
        f'--load-extension={EXTENSION_DIR.resolve()}，'
        f'--disable-extensions-except={EXTENSION_DIR.resolve()}，'
        '--remote-allow-origins=*'
    )
    print(f'profile 已注册扩展路径：{registered_text}')


def ensure_plugin_ready(page: ChromiumPage) -> None:
    for attempt in range(1, PLUGIN_READY_RETRY_TIMES + 1):
        if wait_for_element(page, PLUGIN_ROOT_XPATH, timeout=DEFAULT_TIMEOUT) and wait_for_element(page, PLUGIN_TRIGGER_XPATH, timeout=2):
            return
        print(f'插件入口尚未加载，等待 {PLUGIN_READY_RETRY_INTERVAL} 秒后重试（{attempt}/{PLUGIN_READY_RETRY_TIMES}）...')
        time.sleep(PLUGIN_READY_RETRY_INTERVAL)
    log_extension_injection_failure(page)
    raise RuntimeError(f'扩展未注入：未找到插件入口 {PLUGIN_TRIGGER_XPATH}')


def open_with_cookies(page: ChromiumPage, cookies: list[dict]) -> None:
    # print(f'正在使用固定用户目录启动浏览器：{USER_DATA_DIR}')
    # print(f'请手动安装扩展目录：{EXTENSION_DIR}')
    wait_for_manual_extension_install()

    print(f'正在打开站点首页：{HOME_URL}')
    page.get(HOME_URL, timeout=DEFAULT_TIMEOUT)
    page.wait.doc_loaded(timeout=DEFAULT_TIMEOUT)

    print(f'正在注入 {len(cookies)} 个 cookie...')
    page.set.cookies(cookies)

    print(f'正在打开目标页面：{TARGET_URL}')
    page.get(TARGET_URL, timeout=DEFAULT_TIMEOUT)
    page.wait.doc_loaded(timeout=DEFAULT_TIMEOUT)

    current_path = normalized_path(page.url or '')
    if current_path == '/login':
        raise RuntimeError('cookie 已失效，页面跳回了登录页')


def open_goods_management(page: ChromiumPage) -> None:
    print('正在等待插件入口加载...')
    ensure_plugin_ready(page)

    print('正在打开插件入口...')
    click_element(page, PLUGIN_TRIGGER_XPATH)
    time.sleep(1)

    print('正在进入商品管理...')
    click_element(page, GOODS_MANAGEMENT_XPATH)
    ensure_element(page, SEARCH_INPUT_XPATH, timeout=DEFAULT_TIMEOUT)


def search_goods_batch(page: ChromiumPage, goods_batch: list[str]) -> None:
    query = ' '.join(goods_batch)
    print(f'正在搜索本批商品，共 {len(goods_batch)} 个...')
    input_text(page, SEARCH_INPUT_XPATH, query)
    time.sleep(2)
    ensure_element(page, TABLE_CHECKBOX_LABEL_XPATH, timeout=DEFAULT_TIMEOUT)


def has_search_result_rows(page: ChromiumPage) -> bool:
    rows = page.eles(TABLE_ROW_XPATH, timeout=DEFAULT_TIMEOUT)
    return any((row.attr('data-row-key') or '').strip() for row in rows)


def wait_for_search_result_rows(page: ChromiumPage) -> bool:
    for attempt in range(1, SEARCH_RESULT_RETRY_TIMES + 1):
        if has_search_result_rows(page):
            return True
        print(f'搜索结果尚未出现，等待 {SEARCH_RESULT_RETRY_INTERVAL} 秒后重试（{attempt}/{SEARCH_RESULT_RETRY_TIMES}）...')
        time.sleep(SEARCH_RESULT_RETRY_INTERVAL)
    return has_search_result_rows(page)


def is_end_flash_sale_menu_enabled(page: ChromiumPage) -> bool:
    menu_item = page.ele(END_FLASH_SALE_MENU_XPATH, timeout=DEFAULT_TIMEOUT)
    if not menu_item:
        return False

    class_name = menu_item.attr('class') or ''
    aria_disabled = menu_item.attr('aria-disabled')
    if 'disabled' in class_name or aria_disabled == 'true' or not menu_item.states.is_enabled:
        return False

    return True


def is_checkbox_selectable(page: ChromiumPage) -> bool:
    checkbox_input = page.ele(TABLE_CHECKBOX_INPUT_XPATH, timeout=DEFAULT_TIMEOUT)
    if not checkbox_input:
        return False

    if not checkbox_input.states.is_displayed:
        return False

    disabled = checkbox_input.attr('disabled')
    aria_disabled = checkbox_input.attr('aria-disabled')
    if disabled is not None or aria_disabled == 'true' or not checkbox_input.states.is_enabled:
        return False

    return True


def is_checkbox_checked(page: ChromiumPage) -> bool:
    checkbox_input = page.ele(TABLE_CHECKBOX_INPUT_XPATH, timeout=DEFAULT_TIMEOUT)
    if not checkbox_input:
        return False
    checked = checkbox_input.attr('checked')
    aria_checked = checkbox_input.attr('aria-checked')
    return checked is not None or aria_checked == 'true' or checkbox_input.states.is_checked


def select_all_goods(page: ChromiumPage) -> bool:
    print('正在判断当前批次是否有商品行...')
    if not wait_for_search_result_rows(page):
        print('当前搜索结果没有商品行，跳过本批。')
        return False

    print('正在判断当前批次商品是否可勾选...')
    if not is_checkbox_selectable(page):
        print('当前批次没有可勾选商品，跳过本批。')
        return False

    print('当前批次商品可勾选，开始勾选...')
    for attempt in range(1, SELECT_ALL_RETRY_TIMES + 1):
        checkbox = wait_for_clickable(page, TABLE_CHECKBOX_LABEL_XPATH)
        checkbox.click()
        time.sleep(SELECT_ALL_RETRY_INTERVAL)
        if is_checkbox_checked(page):
            print('当前批次商品已确认勾选。')
            return True
        print(f'勾选状态未生效，等待 {SELECT_ALL_RETRY_INTERVAL} 秒后重试（{attempt}/{SELECT_ALL_RETRY_TIMES}）...')

    print('当前批次商品勾选后仍未生效，跳过本批。')
    return False



def open_bulk_action_menu(page: ChromiumPage, menu_locator: str) -> None:
    if not wait_for_bulk_action_button(page):
        raise RuntimeError('未找到批量操作按钮')
    click_element(page, BULK_ACTION_BUTTON_XPATH)
    if not wait_for_bulk_action_menu_item(page, menu_locator):
        raise RuntimeError(f'未找到批量操作菜单项：{menu_locator}')
    time.sleep(1)


def wait_for_bulk_action_button(page: ChromiumPage) -> bool:
    for attempt in range(1, BULK_ACTION_RETRY_TIMES + 1):
        button = page.ele(BULK_ACTION_BUTTON_XPATH, timeout=DEFAULT_TIMEOUT)
        if button and button.states.is_displayed and button.states.is_enabled:
            return True
        print(f'批量操作按钮尚未出现，等待 {BULK_ACTION_RETRY_INTERVAL} 秒后重试（{attempt}/{BULK_ACTION_RETRY_TIMES}）...')
        time.sleep(BULK_ACTION_RETRY_INTERVAL)
    button = page.ele(BULK_ACTION_BUTTON_XPATH, timeout=DEFAULT_TIMEOUT)
    return bool(button and button.states.is_displayed and button.states.is_enabled)


def wait_for_bulk_action_menu_item(page: ChromiumPage, menu_locator: str) -> bool:
    for attempt in range(1, BULK_ACTION_RETRY_TIMES + 1):
        menu_item = page.ele(menu_locator, timeout=DEFAULT_TIMEOUT)
        if menu_item and menu_item.states.is_displayed:
            return True
        print(f'批量操作菜单尚未出现，等待 {BULK_ACTION_RETRY_INTERVAL} 秒后重试（{attempt}/{BULK_ACTION_RETRY_TIMES}）...')
        time.sleep(BULK_ACTION_RETRY_INTERVAL)
    menu_item = page.ele(menu_locator, timeout=DEFAULT_TIMEOUT)
    return bool(menu_item and menu_item.states.is_displayed)


def end_flash_sale(page: ChromiumPage) -> bool:
    print('正在批量结束限时限量...')
    if not select_all_goods(page):
        return False

    open_bulk_action_menu(page, END_FLASH_SALE_MENU_XPATH)
    print('正在判断“批量结束限时限量”菜单是否可用...')
    if not is_end_flash_sale_menu_enabled(page):
        print('“批量结束限时限量”菜单当前不可用，跳过本批。')
        return False

    click_element(page, END_FLASH_SALE_MENU_XPATH)
    click_element(page, END_FLASH_SALE_CONFIRM_XPATH)
    time.sleep(2)
    return True


def open_bulk_edit(page: ChromiumPage) -> bool:
    print('正在打开批量编辑商品...')
    if not select_all_goods(page):
        return False

    open_bulk_action_menu(page, BULK_EDIT_MENU_XPATH)
    click_element(page, BULK_EDIT_MENU_XPATH)
    ensure_element(page, SUBMIT_BUTTON_XPATH, timeout=DEFAULT_TIMEOUT)
    return True


def prompt_required_text(prompt: str) -> str:
    while True:
        value = input(prompt).strip()
        if value:
            return value
        print('输入不能为空，请重新输入。')


def prompt_required_image_path(prompt: str) -> str:
    while True:
        raw_value = input(prompt).strip().strip('"')
        if not raw_value:
            print('图片路径不能为空，请重新输入。')
            continue
        image_path = Path(raw_value)
        if image_path.exists() and image_path.is_file():
            return str(image_path)
        print('图片不存在，请重新输入。')


def ensure_edit_inputs_provided() -> None:
    global TEXT_TO_FIND, TEXT_TO_REPLACE, IMAGE_PATHS

    if not TEXT_TO_FIND:
        TEXT_TO_FIND = prompt_required_text('请输入原文内容：')
    if not TEXT_TO_REPLACE:
        TEXT_TO_REPLACE = prompt_required_text('请输入新文内容：')

    first_image, second_image = IMAGE_PATHS
    if not first_image:
        first_image = prompt_required_image_path('请输入第 1 张替换图片路径：')
    if not second_image:
        second_image = prompt_required_image_path('请输入第 2 张替换图片路径：')
    IMAGE_PATHS = (first_image, second_image)


def apply_text_replacement(page: ChromiumPage) -> None:
    print('正在执行文本替换...')
    input_text(page, ORIGINAL_TEXT_INPUT_XPATH, TEXT_TO_FIND)
    input_text(page, REPLACEMENT_TEXT_INPUT_XPATH, TEXT_TO_REPLACE)
    click_element(page, TEXT_REPLACE_BUTTON_XPATH)
    time.sleep(1)


def upload_single_image(page: ChromiumPage, trigger_locator: str, image_path: Path) -> None:
    trigger = wait_for_clickable(page, trigger_locator)
    trigger.click.to_upload(str(image_path))
    time.sleep(1)


def apply_image_replacement(page: ChromiumPage) -> None:
    first_path = Path(IMAGE_PATHS[0])
    second_path = Path(IMAGE_PATHS[1])
    if not first_path.exists() or not second_path.exists():
        raise FileNotFoundError('替换图片不存在，请检查图片路径')

    print('正在上传替换图片...')
    upload_single_image(page, IMAGE_UPLOAD_1_XPATH, first_path)
    upload_single_image(page, IMAGE_UPLOAD_2_XPATH, second_path)
    click_element(page, EXECUTE_IMAGE_REPLACE_XPATH)
    time.sleep(2)


def submit_changes(page: ChromiumPage) -> None:
    print('正在提交修改...')
    click_element(page, SUBMIT_BUTTON_XPATH)
    ensure_element(page, REVIEW_DIALOG_XPATH, timeout=DEFAULT_TIMEOUT)


def normalize_status(status_text: str) -> str:
    text = status_text.strip()
    if '通过' in text:
        return '已通过'
    if '失败' in text or '拒绝' in text or '驳回' in text:
        return '失败'
    if '跳过' in text:
        return '跳过'
    return '审核中'




def prompt_concurrency(max_concurrency: int) -> int:
    while True:
        raw_value = input(f'请输入同时执行的组数（1-{max_concurrency}）：').strip()
        if not raw_value.isdigit():
            print('请输入正整数。')
            continue
        concurrency = int(raw_value)
        if concurrency < 1:
            print('并发数不能小于 1。')
            continue
        return min(concurrency, max_concurrency)


def chunk_batches_by_concurrency(goods_batches: list[list[str]], concurrency: int) -> list[list[tuple[int, list[str]]]]:
    indexed_batches = list(enumerate(goods_batches, start=1))
    return [indexed_batches[index:index + concurrency] for index in range(0, len(indexed_batches), concurrency)]


def initialize_result_file(output_file: Path) -> None:
    save_results([], output_file)


def append_results(results: list[dict[str, str]], output_file: Path) -> None:
    existing_results: list[dict[str, str]] = []
    if output_file.exists():
        try:
            payload = json.loads(output_file.read_text(encoding='utf-8'))
            loaded_results = payload.get('results', [])
            if isinstance(loaded_results, list):
                existing_results = loaded_results
        except json.JSONDecodeError:
            existing_results = []

    merged_results = [*existing_results, *results]
    save_results(merged_results, output_file)


def parse_progress_percent(dialog) -> int | None:
    progress = dialog.ele('xpath:.//span[contains(@class, "sp-pct")]', timeout=2)
    if not progress:
        return None
    text = progress.text.strip().rstrip('%')
    if not text.isdigit():
        return None
    return int(text)


def parse_status_cards(page: ChromiumPage, dialog_locator: str) -> list[dict[str, str]]:
    dialog = ensure_element(page, dialog_locator, timeout=DEFAULT_TIMEOUT)
    cards = dialog.eles(REVIEW_CARD_XPATH)
    results: list[dict[str, str]] = []

    for card in cards:
        name_element = card.ele(REVIEW_NAME_XPATH, timeout=2)
        id_element = card.ele(REVIEW_ID_XPATH, timeout=2)
        status_element = card.ele(REVIEW_STATUS_XPATH, timeout=2)
        if not name_element or not id_element or not status_element:
            continue

        raw_id_text = id_element.text.strip()
        goods_id = raw_id_text.split('ID:')[-1].strip() if 'ID:' in raw_id_text else raw_id_text
        results.append(
            {
                'goods_id': goods_id,
                'goods_name': name_element.text.strip(),
                'review_status': normalize_status(status_element.text),
            }
        )

    return results


def all_reviews_finished(results: list[dict[str, str]]) -> bool:
    return bool(results) and all(item['review_status'] != '审核中' for item in results)


def collect_dialog_results(page: ChromiumPage, dialog_locator: str, timeout: int | None = None) -> list[dict[str, str]]:
    timeout = REVIEW_TIMEOUT if timeout is None else timeout
    print(f'正在采集处理结果，最长等待 {timeout} 秒...')
    end_time = time.time() + timeout
    latest_results: list[dict[str, str]] = []

    while time.time() < end_time:
        dialog = ensure_element(page, dialog_locator, timeout=DEFAULT_TIMEOUT)
        latest_results = parse_status_cards(page, dialog_locator)
        progress_percent = parse_progress_percent(dialog)
        if progress_percent == 100 or all_reviews_finished(latest_results):
            return latest_results
        time.sleep(POLL_INTERVAL)

    return latest_results


def collect_review_results(page: ChromiumPage, timeout: int | None = None) -> list[dict[str, str]]:
    return collect_dialog_results(page, REVIEW_DIALOG_XPATH, timeout=timeout)


def collect_restore_results(page: ChromiumPage, timeout: int | None = None) -> list[dict[str, str]]:
    return collect_dialog_results(page, RESTORE_DIALOG_XPATH, timeout=timeout)


def close_review_dialog(page: ChromiumPage) -> None:
    print('正在关闭审核结果弹窗...')
    if wait_for_element(page, REVIEW_CLOSE_BUTTON_XPATH, timeout=DEFAULT_TIMEOUT):
        click_element(page, REVIEW_CLOSE_BUTTON_XPATH)
        time.sleep(1)
        return
    click_element_by_text_contains(page, 'button', '关 闭', timeout=DEFAULT_TIMEOUT)
    time.sleep(1)


def filter_failed_goods_ids(results: list[dict[str, str]]) -> list[str]:
    return [item['goods_id'] for item in results if item['review_status'] != '已通过']


def build_goods_row_locator(goods_id: str) -> str:
    return f'xpath://tr[@data-row-key="{goods_id}"]'


def build_goods_row_checkbox_label_locator(goods_id: str) -> str:
    return f'xpath://tr[@data-row-key="{goods_id}"]/td[1]//label'


def build_goods_row_checkbox_input_locator(goods_id: str) -> str:
    return f'xpath://tr[@data-row-key="{goods_id}"]/td[1]//input'


def is_goods_row_checked(page: ChromiumPage, goods_id: str) -> bool:
    checkbox_input = page.ele(build_goods_row_checkbox_input_locator(goods_id), timeout=DEFAULT_TIMEOUT)
    if not checkbox_input:
        return False
    checked = checkbox_input.attr('checked')
    aria_checked = checkbox_input.attr('aria-checked')
    return checked is not None or aria_checked == 'true' or checkbox_input.states.is_checked


def wait_for_goods_table_ready(page: ChromiumPage) -> None:
    print('正在等待商品页面重新加载...')
    ensure_element(page, SEARCH_INPUT_XPATH, timeout=DEFAULT_TIMEOUT)
    ensure_element(page, TABLE_CHECKBOX_CONTAINER_XPATH, timeout=DEFAULT_TIMEOUT)
    ensure_element(page, TABLE_CHECKBOX_LABEL_XPATH, timeout=DEFAULT_TIMEOUT)
    if not wait_for_search_result_rows(page):
        raise RuntimeError('商品页面未重新加载出商品行')


def wait_for_restore_dialog(page: ChromiumPage) -> None:
    for attempt in range(1, RESTORE_DIALOG_RETRY_TIMES + 1):
        if wait_for_element(page, RESTORE_DIALOG_XPATH, timeout=DEFAULT_TIMEOUT):
            return
        print(f'恢复结果弹窗尚未出现，等待 {RESTORE_DIALOG_RETRY_INTERVAL} 秒后重试（{attempt}/{RESTORE_DIALOG_RETRY_TIMES}）...')
        time.sleep(RESTORE_DIALOG_RETRY_INTERVAL)
    raise RuntimeError(f'未找到元素：{RESTORE_DIALOG_XPATH}')


def ensure_goods_row_checked(page: ChromiumPage, goods_id: str) -> None:
    ensure_element(page, build_goods_row_locator(goods_id), timeout=DEFAULT_TIMEOUT)
    if is_goods_row_checked(page, goods_id):
        return

    for attempt in range(1, SELECT_ALL_RETRY_TIMES + 1):
        click_element(page, build_goods_row_checkbox_label_locator(goods_id), timeout=DEFAULT_TIMEOUT)
        time.sleep(SELECT_ALL_RETRY_INTERVAL)
        if is_goods_row_checked(page, goods_id):
            return
        print(f'商品 {goods_id} 勾选状态未生效，等待 {SELECT_ALL_RETRY_INTERVAL} 秒后重试（{attempt}/{SELECT_ALL_RETRY_TIMES}）...')
    raise RuntimeError(f'未能勾选商品：{goods_id}')


def deselect_goods_ids(page: ChromiumPage, goods_ids: list[str]) -> None:
    if not goods_ids:
        return

    print(f'正在取消勾选未通过商品，共 {len(goods_ids)} 个...')
    for goods_id in goods_ids:
        ensure_element(page, build_goods_row_locator(goods_id), timeout=DEFAULT_TIMEOUT)
        if not is_goods_row_checked(page, goods_id):
            ensure_goods_row_checked(page, goods_id)

        for attempt in range(1, SELECT_ALL_RETRY_TIMES + 1):
            click_element(page, build_goods_row_checkbox_label_locator(goods_id), timeout=DEFAULT_TIMEOUT)
            time.sleep(SELECT_ALL_RETRY_INTERVAL)
            if not is_goods_row_checked(page, goods_id):
                break
            print(f'商品 {goods_id} 取消勾选未生效，等待 {SELECT_ALL_RETRY_INTERVAL} 秒后重试（{attempt}/{SELECT_ALL_RETRY_TIMES}）...')
        if is_goods_row_checked(page, goods_id):
            raise RuntimeError(f'未能取消勾选商品：{goods_id}')


def open_flash_sale_panel(page: ChromiumPage) -> None:
    print('正在打开限时限量面板...')
    if wait_for_element(page, FLASH_SALE_BUTTON_XPATH, timeout=DEFAULT_TIMEOUT):
        click_element(page, FLASH_SALE_BUTTON_XPATH)
    else:
        click_element_by_text_contains(page, 'button', '限时限量', timeout=DEFAULT_TIMEOUT)
    ensure_element(page, FLASH_SALE_AMOUNT_INPUT_XPATH, timeout=DEFAULT_TIMEOUT)


def create_flash_sale(page: ChromiumPage, amount: str | None = None) -> None:
    amount = FLASH_SALE_AMOUNT if amount is None else amount
    open_flash_sale_panel(page)
    print(f'正在输入限时限量金额：{amount}...')
    input_text(page, FLASH_SALE_AMOUNT_INPUT_XPATH, amount)
    if wait_for_element(page, FLASH_SALE_APPLY_BUTTON_XPATH, timeout=DEFAULT_TIMEOUT):
        click_element(page, FLASH_SALE_APPLY_BUTTON_XPATH)
    else:
        click_element_by_text_contains(page, 'button', '应用到选中', timeout=DEFAULT_TIMEOUT)
    time.sleep(1)
    if wait_for_element(page, FLASH_SALE_CREATE_BUTTON_XPATH, timeout=DEFAULT_TIMEOUT):
        click_element(page, FLASH_SALE_CREATE_BUTTON_XPATH)
    else:
        click_element_by_text_contains(page, 'button', '一键创建', timeout=DEFAULT_TIMEOUT)
    time.sleep(2)


def restore_selection(page: ChromiumPage, failed_goods_ids: list[str]) -> list[dict[str, str]]:
    wait_for_goods_table_ready(page)
    if not select_all_goods(page):
        raise RuntimeError('恢复阶段重新勾选商品失败')
    deselect_goods_ids(page, failed_goods_ids)

    print('正在恢复选中商品...')
    if wait_for_element(page, RESTORE_SELECTION_BUTTON_XPATH, timeout=DEFAULT_TIMEOUT):
        click_element(page, RESTORE_SELECTION_BUTTON_XPATH)
    else:
        click_element_by_text_contains(page, 'button', '恢复选中', timeout=DEFAULT_TIMEOUT)
    wait_for_restore_dialog(page)
    return collect_restore_results(page)


def run_post_review_flow(page: ChromiumPage, review_results: list[dict[str, str]]) -> list[dict[str, str]]:
    failed_goods_ids = filter_failed_goods_ids(review_results)
    passed_count = len(review_results) - len(failed_goods_ids)

    close_review_dialog(page)
    wait_for_goods_table_ready(page)

    if passed_count <= 0:
        print('本批没有审核通过的商品，跳过后续限时限量与恢复流程。')
        return []

    if not select_all_goods(page):
        raise RuntimeError('审核后重新勾选商品失败')
    deselect_goods_ids(page, failed_goods_ids)
    create_flash_sale(page)
    return restore_selection(page, failed_goods_ids)


def build_batch_skip_result(goods_batch: list[str], reason: str) -> dict:
    return {
        'status': 'skipped',
        'reason': reason,
        'goods_ids': goods_batch,
        'results': [],
        'review_results': [],
    }


def build_batch_failed_result(
    goods_batch: list[str],
    reason: str,
    review_results: list[dict[str, str]] | None = None,
) -> dict:
    return {
        'status': 'failed',
        'reason': reason,
        'goods_ids': goods_batch,
        'results': [],
        'review_results': review_results or [],
    }


def build_batch_success_result(
    goods_batch: list[str],
    results: list[dict[str, str]],
    review_results: list[dict[str, str]],
) -> dict:
    return {
        'status': 'completed',
        'reason': None,
        'goods_ids': goods_batch,
        'results': results,
        'review_results': review_results,
    }


def print_batch_status_result(
    batch_index: int,
    batch_count: int,
    wave_index: int,
    wave_count: int,
    worker_index: int,
    batch_result: dict,
) -> None:
    print(
        json.dumps(
            {
                'wave_index': wave_index,
                'wave_count': wave_count,
                'worker_index': worker_index,
                'batch_index': batch_index,
                'batch_count': batch_count,
                'status': batch_result['status'],
                'reason': batch_result['reason'],
                'goods_ids': batch_result['goods_ids'],
            },
            ensure_ascii=False,
        )
    )


def print_review_results(results: list[dict[str, str]], title: str = '审核结果如下：') -> None:
    print(title)
    if not results:
        print('未读取到审核结果。')
        return

    for item in results:
        print(
            json.dumps(
                {
                    'goods_id': item['goods_id'],
                    'goods_name': item['goods_name'],
                    'review_status': item['review_status'],
                },
                ensure_ascii=False,
            )
        )


def save_results(results: list[dict[str, str]], output_file: Path) -> None:
    payload = {
        'saved_at': time.strftime('%Y-%m-%dT%H:%M:%S'),
        'count': len(results),
        'results': results,
    }
    output_file.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding='utf-8')
    print(f'结果已保存到：{output_file.resolve()}')


def save_review_results(results: list[dict[str, str]], output_file: Path | None = None) -> None:
    output_file = REVIEW_RESULTS_FILE if output_file is None else output_file
    save_results(results, output_file)


def append_review_results(results: list[dict[str, str]], output_file: Path | None = None) -> None:
    output_file = REVIEW_RESULTS_FILE if output_file is None else output_file
    append_results(results, output_file)


def process_goods_batch(page: ChromiumPage, goods_batch: list[str]) -> dict:
    search_goods_batch(page, goods_batch)

    if not wait_for_search_result_rows(page):
        return build_batch_skip_result(goods_batch, '当前搜索结果没有商品行')

    if not is_checkbox_selectable(page):
        return build_batch_skip_result(goods_batch, '当前批次没有可勾选商品')

    print('正在批量结束限时限量...')
    if not select_all_goods(page):
        return build_batch_skip_result(goods_batch, '当前批次商品勾选失败')
    open_bulk_action_menu(page, END_FLASH_SALE_MENU_XPATH)
    print('正在判断“批量结束限时限量”菜单是否可用...')
    if not is_end_flash_sale_menu_enabled(page):
        return build_batch_skip_result(goods_batch, '批量结束限时限量菜单当前不可用')

    click_element(page, END_FLASH_SALE_MENU_XPATH)
    click_element(page, END_FLASH_SALE_CONFIRM_XPATH)
    time.sleep(2)

    print('正在打开批量编辑商品...')
    if not wait_for_search_result_rows(page):
        return build_batch_skip_result(goods_batch, '结束限时限量后当前结果区没有商品行')

    if not is_checkbox_selectable(page):
        return build_batch_skip_result(goods_batch, '结束限时限量后商品不可再次勾选')

    if not select_all_goods(page):
        return build_batch_skip_result(goods_batch, '结束限时限量后重新勾选失败')
    open_bulk_action_menu(page, BULK_EDIT_MENU_XPATH)
    click_element(page, BULK_EDIT_MENU_XPATH)
    ensure_element(page, SUBMIT_BUTTON_XPATH, timeout=DEFAULT_TIMEOUT)

    apply_text_replacement(page)
    apply_image_replacement(page)
    submit_changes(page)
    review_results = collect_review_results(page)

    try:
        final_results = run_post_review_flow(page, review_results)
    except Exception as exc:
        return build_batch_failed_result(goods_batch, f'审核后续流程失败：{exc}', review_results=review_results)

    return build_batch_success_result(goods_batch, final_results, review_results)


def run_batch_worker(tab, cookies: list[dict], goods_batch: list[str]) -> dict:
    try:
        open_with_cookies(tab, cookies)
        open_goods_management(tab)
        return process_goods_batch(tab, goods_batch)
    finally:
        tab.close()


def execute_wave(
    root_page: ChromiumPage,
    cookies: list[dict],
    wave_batches: list[tuple[int, list[str]]],
    wave_index: int,
    wave_count: int,
    batch_count: int,
    review_results_file: Path,
    restore_results_file: Path,
) -> tuple[list[dict[str, str]], list[dict[str, str]], list[dict]]:
    wave_review_results: list[dict[str, str]] = []
    wave_restore_results: list[dict[str, str]] = []
    wave_batch_results: list[dict] = []
    futures: dict = {}

    with ThreadPoolExecutor(max_workers=len(wave_batches)) as executor:
        for worker_index, (batch_index, goods_batch) in enumerate(wave_batches, start=1):
            print(f'第 {wave_index}/{wave_count} 波启动 worker {worker_index}，处理第 {batch_index}/{batch_count} 批。')
            tab = root_page.new_tab('about:blank')
            future = executor.submit(run_batch_worker, tab, cookies, goods_batch)
            futures[future] = {
                'worker_index': worker_index,
                'batch_index': batch_index,
                'goods_batch': goods_batch,
            }

        for future in as_completed(futures):
            metadata = futures[future]
            worker_index = metadata['worker_index']
            batch_index = metadata['batch_index']
            goods_batch = metadata['goods_batch']

            try:
                batch_result = future.result()
            except Exception as exc:
                batch_result = build_batch_failed_result(goods_batch, str(exc))

            print_batch_status_result(
                batch_index=batch_index,
                batch_count=batch_count,
                wave_index=wave_index,
                wave_count=wave_count,
                worker_index=worker_index,
                batch_result=batch_result,
            )
            wave_batch_results.append(batch_result)

            review_results = batch_result.get('review_results', [])
            final_results = batch_result.get('results', [])
            if review_results:
                append_results(review_results, review_results_file)
                wave_review_results.extend(review_results)
            if final_results:
                append_results(final_results, restore_results_file)
                wave_restore_results.extend(final_results)

    return wave_review_results, wave_restore_results, wave_batch_results


def load_config_file(config_file: Path) -> AutomationConfig:
    payload = json.loads(config_file.read_text(encoding='utf-8'))
    if not isinstance(payload, dict):
        raise ValueError(f'配置文件格式错误：{config_file}')
    return AutomationConfig.from_dict(payload)


def _run_automation(config: AutomationConfig) -> tuple[list[dict[str, str]], list[dict[str, str]]]:
    apply_runtime_config(config)
    log_startup_diagnostics()
    cookies = load_cookies(config.cookie_file)
    excel_file = config.excel_file or resolve_excel_file()
    goods_ids = load_goods_ids_from_excel(excel_file)
    goods_batches = chunk_goods_ids(goods_ids, config.max_batch_size)
    review_results_file = config.review_results_file
    restore_results_file = config.restore_results_file
    concurrency = min(max(1, config.concurrency), len(goods_batches))
    goods_batch_waves = chunk_batches_by_concurrency(goods_batches, concurrency)
    all_results: list[dict[str, str]] = []
    all_review_results: list[dict[str, str]] = []
    failed_batches: list[dict] = []

    review_results_file.parent.mkdir(parents=True, exist_ok=True)
    restore_results_file.parent.mkdir(parents=True, exist_ok=True)

    print(f'已读取 Excel 文件：{excel_file}')
    print(f'共读取到 {len(goods_ids)} 个商品 ID，将分为 {len(goods_batches)} 批处理。')
    print(f'将按 {concurrency} 个并发 worker 分波执行，共 {len(goods_batch_waves)} 波。')

    ensure_edit_inputs_provided()
    initialize_result_file(review_results_file)
    initialize_result_file(restore_results_file)

    root_page = build_page()
    try:
        for wave_index, wave_batches in enumerate(goods_batch_waves, start=1):
            print(f'开始执行第 {wave_index}/{len(goods_batch_waves)} 波，本波共 {len(wave_batches)} 批。')
            wave_review_results, wave_restore_results, wave_batch_results = execute_wave(
                root_page=root_page,
                cookies=cookies,
                wave_batches=wave_batches,
                wave_index=wave_index,
                wave_count=len(goods_batch_waves),
                batch_count=len(goods_batches),
                review_results_file=review_results_file,
                restore_results_file=restore_results_file,
            )
            all_review_results.extend(wave_review_results)
            all_results.extend(wave_restore_results)
            failed_batches.extend(
                batch_result for batch_result in wave_batch_results if batch_result.get('status') == 'failed'
            )

        print_review_results(all_review_results, title='提交审核结果如下：')
        print_review_results(all_results, title='最终恢复结果如下：')
        if failed_batches:
            first_error = failed_batches[0].get('reason') or failed_batches[0].get('message') or '未知错误'
            raise RuntimeError(f'有 {len(failed_batches)} 个批次失败，任务标记为失败。首个错误：{first_error}')
        return all_review_results, all_results
    finally:
        root_page.quit()
        print('浏览器已关闭。')


def run_automation(
    config: AutomationConfig,
    log_callback: Callable[[str], None] | None = None,
) -> tuple[list[dict[str, str]], list[dict[str, str]]]:
    if log_callback is None:
        return _run_automation(config)

    writer = CallbackWriter(log_callback)
    with contextlib.redirect_stdout(writer), contextlib.redirect_stderr(writer):
        try:
            return _run_automation(config)
        finally:
            writer.flush()


def build_interactive_config() -> AutomationConfig:
    excel_file = resolve_excel_file()
    goods_ids = load_goods_ids_from_excel(excel_file)
    goods_batches = chunk_goods_ids(goods_ids, MAX_BATCH_SIZE)
    concurrency = prompt_concurrency(len(goods_batches))
    ensure_edit_inputs_provided()
    return AutomationConfig(
        excel_file=excel_file,
        concurrency=concurrency,
        text_to_find=TEXT_TO_FIND,
        text_to_replace=TEXT_TO_REPLACE,
        image_paths=(Path(IMAGE_PATHS[0]), Path(IMAGE_PATHS[1])),
    )


def main() -> None:
    parser = argparse.ArgumentParser(description='拼多多商品批量自动化')
    parser.add_argument('--config', help='从 JSON 配置文件读取自动化参数')
    args = parser.parse_args()
    config = load_config_file(Path(args.config)) if args.config else build_interactive_config()
    run_automation(config)


if __name__ == '__main__':
    main()
